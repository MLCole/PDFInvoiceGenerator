from fileinput import filename
import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from pathlib import Path

# Get list of files
filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    #print(df)
    # Create 1 PDF Document per excel file
    pdf = FPDF(orientation="P", unit="mm", format="A4") # 'P' stands for Portrait, 'L' for Landscape
    pdf.add_page()
    ## Get filename without the extension
    filename = Path(filepath).stem
    invoice_num, invoice_date = filename.split("-")
    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, text=f"Invoice #: {invoice_num}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(w=50, h=8, text=f"Date: {invoice_date}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)


    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    # Add your headers
    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, text=str(columns[0]), border=1)
    pdf.cell(w=70, h=8, text=str(columns[1]), border=1)
    pdf.cell(w=32, h=8, text=str(columns[2]), border=1)
    pdf.cell(w=30, h=8, text=str(columns[3]), border=1)
    pdf.cell(w=30, h=8, text=str(columns[4]), border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Add rows to table.
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, text=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, text=str(row["product_name"]), border=1)
        pdf.cell(w=32, h=8, text=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, text=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, text=str(row["total_price"]), border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    total_sum = df["total_price"].sum()

    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, text="", border=1)
    pdf.cell(w=70, h=8, text="", border=1)
    pdf.cell(w=32, h=8, text="", border=1)
    pdf.cell(w=30, h=8, text="", border=1)
    pdf.cell(w=30, h=8, text=f"${str(total_sum)}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(10)


    # Add total sum information
    pdf.set_font(family="Times", size=14, style="B")
    pdf.cell(w=0, h=8, text=f"The total due is ${str(total_sum)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Add Company Name & Logo
    pdf.set_font(family="Times", size=14, style="B")
    pdf.cell(w=88, h=8, text=f"Pay to the order of: SuperAwesomeTeam")
    pdf.image("pythonhow.png", w=10)

    # Thank them!
    pdf.ln(50)
    pdf.set_font(family="Times", size=18, style="B")
    pdf.cell(w=0, h=8, text=f"                                               Thank you so much for your business!", new_x=XPos.LMARGIN, new_y=YPos.NEXT)


    pdf.output(f"PDFs/{filename}.pdf")